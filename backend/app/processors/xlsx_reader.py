from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

def read_xlsx_rows(path: str | Path) -> list[dict[str, object]]:
    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(value).strip() if value is not None else "" for value in headers_row]
    normalized_headers = [header or f"column_{index + 1}" for index, header in enumerate(headers)]

    rows: list[dict[str, object]] = []
    for values in rows_iter:
        row = {normalized_headers[index]: values[index] for index in range(len(normalized_headers))}
        rows.append(row)

    workbook.close()
    return rows
